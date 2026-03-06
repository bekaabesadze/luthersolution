"""
camel_excel_parser.py - Parse WSB Performance Dashboard style Excel workbooks.

Extracts CAMELS strategic metrics from sheets like "FY24", "FY 25".
Expects: row 1 = title (bank name), row 2 = headers (CAMELS, Strategic Metrics, Q1 24 WSB, ...),
data rows with column A = category, B = metric name, then value columns per quarter.
"""

import re
from typing import List, Optional, Tuple

import openpyxl


# Map Excel metric labels to canonical names (stored in DB; must match frontend keys).
EXCEL_TO_CANONICAL = {
    "Equity to Assets": "Equity to Assets",
    "Tier 1 Leverage Ratio": "Tier 1 Leverage Ratio",
    "Legal Lending Limit": "Legal Lending Limit",
    "90+ PD and Non-Accrual / Total Loans": "90+ PD and Non-Accrual / Total Loans",
    "Past Due 30-89 / Total Loans": "Past Due 30-89 / Total Loans",
    "Non Performing Assets to Equity + Reserves": "Non Performing Assets to Equity + Reserves",
    "ACL / Total Loans": "ACL / Total Loans",
    "Percentage on Budget": "Percentage on Budget",
    "Number of Employees": "Number of Employees",
    "Efficency Ratio": "Efficiency Ratio",
    "Efficiency Ratio": "Efficiency Ratio",
    "Loan Growth Rate": "Loan Growth Rate",
    "Desposit Growth Rate": "Deposit Growth Rate",
    "Deposit Growth Rate": "Deposit Growth Rate",
    "Non-Interest Income": "Non-Interest Income",
    "Cost of Funds": "Cost of Funds",
    "Yield on Loans": "Yield on Loans",
    "TE Yield on Securities": "TE Yield on Securities",
    "Net Interest Margin": "Net Interest Margin",
    "Return on Assets": "Return on Assets",
    "TE Return on Assets": "Return on Assets",
    "Return on Equity": "Return on Equity",
    "TE Return on Equity": "Return on Equity",
    "Loan to Deposit Ratio": "Loan to Deposit Ratio",
    "Brokered Deposits to Total Deposits": "Brokered Deposits to Total Deposits",
    "Short Term Non-Core Funding": "Short Term Non-Core Funding",
    "Net Loans to Average Assets": "Net Loans to Average Assets",
    "Core Deposits to Average Assets": "Core Deposits to Average Assets",
    "Pledged Assets to Total Assets": "Pledged Assets to Total Assets",
    "FHLB Open Borrowing Capacity": "FHLB Open Borrowing Capacity",
    "Earnings at Risk (12m) -100": "Earnings at Risk (12m) -100",
    "Earnings at Risk (12m) +100": "Earnings at Risk (12m) +100",
    "Economic Value of Equity Ratio -100": "Economic Value of Equity Ratio -100",
    "Economic Value of Equity Ratio +100": "Economic Value of Equity Ratio +100",
    "Non Parallel EAR (12m) Most Likely": "Non Parallel EAR (12m) Most Likely",
    "Non Parallel EVE Ratio Most Likely": "Non Parallel EVE Ratio Most Likely",
}

# Header pattern: "Q1 24 WSB", "Q4 25 WSB", "Q2 24 WSB" -> (year 2024/2025, quarter 1-4)
QUARTER_HEADER = re.compile(r"Q([1-4])\s+(\d{2})\s+(\w+)", re.I)


def _parse_sheet_year(sheet_name: str) -> Optional[int]:
    """Parse fiscal year from sheet name: FY24, FY 25 -> 2024, 2025."""
    m = re.match(r"FY\s*(\d{2,4})\s*$", sheet_name.strip(), re.I)
    if not m:
        return None
    y = int(m.group(1))
    if y < 100:
        y += 2000  # 24 -> 2024
    return y


def _bank_column_headers(row: tuple) -> List[Tuple[int, int, int, str]]:
    """
    From header row (row index 1), return list of (col_index, year, quarter, bank_suffix).
    Only columns that look like "Qn YY BankName" (e.g. Q4 25 WSB).
    """
    result = []
    for col_idx, cell in enumerate(row):
        if cell is None:
            continue
        s = str(cell).strip()
        m = QUARTER_HEADER.search(s)
        if not m:
            continue
        q = int(m.group(1))
        yy = int(m.group(2))
        bank_suffix = (m.group(3) or "").strip()
        year = 2000 + yy if yy < 100 else yy
        result.append((col_idx, year, q, bank_suffix))
    return result


def _extract_bank_name(title_cell) -> str:
    """From row 1 first cell, e.g. 'Waukon State Bank Scoreboard' -> 'Waukon State Bank'."""
    if title_cell is None:
        return "Bank"
    s = str(title_cell).strip()
    for suffix in (" Scoreboard", " Dashboard", " Performance Dashboard"):
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
            break
    return s or "Bank"


def _numeric_value(val) -> Optional[float]:
    """Return float if val is numeric, else None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def parse_camel_excel(
    path: str,
    *,
    bank_name_override: Optional[str] = None,
) -> List[dict]:
    """
    Parse a WSB Performance Dashboard style Excel file.

    Returns list of dicts: {"bank_id", "year", "quarter", "metric_name", "value"}.
    metric_name is canonical (e.g. "Efficiency Ratio", "Deposit Growth Rate").
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out = []

    try:
        for sheet_name in wb.sheetnames:
            sheet_year = _parse_sheet_year(sheet_name)
            if sheet_year is None:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=1, max_row=70, values_only=True))
            if len(rows) < 3:
                continue

            title_row = rows[0]
            header_row = rows[1]
            bank_name = (bank_name_override or _extract_bank_name(title_row[0] if title_row else None)).strip() or "Bank"

            # Columns that are "Qn YY <Bank>" - we take all that have the same bank suffix (e.g. WSB)
            bank_cols = _bank_column_headers(header_row)
            if not bank_cols:
                continue

            # Use first bank suffix found (e.g. WSB) and take all columns for that suffix
            first_suffix = bank_cols[0][3]
            cols_for_bank = [(ci, yr, q) for ci, yr, q, suf in bank_cols if suf == first_suffix]

            for row in rows[2:]:
                metric_raw = row[1] if len(row) > 1 else None
                if metric_raw is None or str(metric_raw).strip() == "":
                    continue
                metric_str = str(metric_raw).strip()
                # Stop at non-metric rows (e.g. "Peer Group" starts another section)
                if metric_str == "Peer Group":
                    break
                canonical = EXCEL_TO_CANONICAL.get(metric_str, metric_str)

                for col_idx, year, quarter in cols_for_bank:
                    if col_idx >= len(row):
                        continue
                    val = _numeric_value(row[col_idx])
                    if val is None:
                        continue
                    rows_out.append({
                        "bank_id": bank_name,
                        "year": year,
                        "quarter": quarter,
                        "metric_name": canonical,
                        "value": val,
                    })
    finally:
        wb.close()

    return rows_out


def parse_camel_excel_from_filelike(file_handle, *, bank_name_override: Optional[str] = None) -> List[dict]:
    """
    Same as parse_camel_excel but from an open file-like object (e.g. UploadFile.file).
    Writes to a temp file then parses, since openpyxl needs a path or bytes.
    """
    import tempfile
    import os
    content = file_handle.read()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        try:
            tmp.write(content)
            tmp.flush()
            return parse_camel_excel(tmp.name, bank_name_override=bank_name_override)
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass
